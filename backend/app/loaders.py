"""문서 형식별 로더 — PDF, DOCX, XLSX, HWP, TXT, MD, HTML, CSV"""

import os
import csv
import zipfile
import io
import xml.etree.ElementTree as ET
from langchain_core.documents import Document
from langchain_community.document_loaders import PyPDFLoader


def load_pdf(file_path: str) -> list[Document]:
    loader = PyPDFLoader(file_path)
    return loader.load()


def load_docx(file_path: str) -> list[Document]:
    import docx2txt
    text = docx2txt.process(file_path)
    return [Document(page_content=text, metadata={"page": 0})]


def load_xlsx(file_path: str) -> list[Document]:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    docs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            text = f"[시트: {sheet_name}]\n" + "\n".join(rows)
            docs.append(Document(page_content=text, metadata={"page": 0, "sheet": sheet_name}))
    wb.close()
    return docs


def load_hwpx(file_path: str) -> list[Document]:
    """HWPX (한/글 XML 형식) 로더 — ZIP 내부 section XML에서 텍스트 추출."""
    texts = []
    with zipfile.ZipFile(file_path, "r") as zf:
        section_files = sorted(
            [n for n in zf.namelist() if n.startswith("Contents/section") and n.endswith(".xml")]
        )
        for section_file in section_files:
            with zf.open(section_file) as f:
                tree = ET.parse(f)
                root = tree.getroot()
                # 모든 텍스트 노드 추출
                for elem in root.iter():
                    if elem.text and elem.text.strip():
                        texts.append(elem.text.strip())
    content = "\n".join(texts) if texts else ""
    return [Document(page_content=content, metadata={"page": 0})] if content else []


def load_hwp(file_path: str) -> list[Document]:
    """HWP (한/글 바이너리 형식) 로더 — olefile로 텍스트 스트림 추출."""
    import olefile
    ole = olefile.OleFileIO(file_path)
    text = ""
    if ole.exists("PrvText"):
        raw = ole.openstream("PrvText").read()
        text = raw.decode("utf-16-le", errors="ignore").strip()
    elif ole.exists("BodyText/Section0"):
        raw = ole.openstream("BodyText/Section0").read()
        # 바이너리에서 텍스트 추출 (널 바이트 기준 분리)
        decoded = raw.decode("utf-16-le", errors="ignore")
        # 제어 문자 제거, 출력 가능 문자만 유지
        text = "".join(c if c.isprintable() or c in "\n\t" else " " for c in decoded).strip()
    ole.close()
    return [Document(page_content=text, metadata={"page": 0})] if text else []


def load_text(file_path: str) -> list[Document]:
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()
    return [Document(page_content=text, metadata={"page": 0})]


def load_html(file_path: str) -> list[Document]:
    from bs4 import BeautifulSoup
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        soup = BeautifulSoup(f.read(), "html.parser")
    text = soup.get_text(separator="\n", strip=True)
    return [Document(page_content=text, metadata={"page": 0})]


def load_csv(file_path: str) -> list[Document]:
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        rows = ["\t".join(row) for row in reader if any(row)]
    text = "\n".join(rows)
    return [Document(page_content=text, metadata={"page": 0})]


# 확장자 → 로더 매핑
LOADER_MAP = {
    ".pdf": load_pdf,
    ".docx": load_docx,
    ".doc": load_docx,
    ".xlsx": load_xlsx,
    ".xls": load_xlsx,
    ".hwp": load_hwp,
    ".hwpx": load_hwpx,
    ".txt": load_text,
    ".md": load_text,
    ".html": load_html,
    ".htm": load_html,
    ".csv": load_csv,
}

SUPPORTED_EXTENSIONS = set(LOADER_MAP.keys())


def load_document(file_path: str) -> list[Document]:
    """파일 확장자에 따라 적절한 로더를 선택하여 문서를 로드한다."""
    ext = os.path.splitext(file_path)[1].lower()
    loader_fn = LOADER_MAP.get(ext)
    if loader_fn is None:
        raise ValueError(f"지원하지 않는 파일 형식: {ext}")
    return loader_fn(file_path)
